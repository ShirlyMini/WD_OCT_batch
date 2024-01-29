# Read data from excel
# Write data to excel
# Data-driven test case

# pip install openpyxl, pandas, xlwt, xlrt

#xl--> workbook-->sheet--->row and col
# row and col num start from 1 to n+1

# import openpyxl
#
# wb=openpyxl.load_workbook(r"C:\Users\user\PycharmProjects\pythonProject_WDBatch\selenium\day35\sample.xlsx")
# print(wb)#<openpyxl.workbook.workbook.Workbook object at 0x0000022D9DD5DF10>
#
# print(wb.sheetnames)#['Sheet1', 'Sheet2', 'Sheet3']
# sheet = wb["Sheet1"]
# print(sheet.max_row)
# print(sheet.max_column)
#
# # print(sheet.cell(3,3).value)
#
# ##################reading####################
#
# for r in range(1, sheet.max_row+1):#(1,11+1)
#     for c in range(1, sheet.max_column+1):
#         print(sheet.cell(r, c).value, end='\t')
#     print("\n")


################################################################################
# import openpyxl
#
# wb=openpyxl.load_workbook(r"C:\Users\user\PycharmProjects\pythonProject_WDBatch\selenium\day35\sample.xlsx")
# sheet = wb["Sheet2"]
#
# # fix row and colo
#
# row_n=6
# col_n=6
#
# for r in range(1, row_n+1):#(1,11+1)
#     for c in range(1, col_n+1):
#         sheet.cell(r, c).value="data"
#
# wb.save(r"C:\Users\user\PycharmProjects\pythonProject_WDBatch\selenium\day35\sample.xlsx")



#
# import openpyxl
# file=r"C:\Users\user\PycharmProjects\pythonProject_WDBatch\selenium\day35\sample.xlsx"
# wb=openpyxl.load_workbook(file)
# sheet = wb["Sheet3"]
#
# sheet.cell(1, 1).value="id"
# sheet.cell(1, 2).value="name"
# sheet.cell(1, 3).value="mark"
#
# sheet.cell(2, 1).value="123"
# sheet.cell(2, 2).value="sita"
# sheet.cell(2, 3).value="100"
#
# sheet.cell(3, 1).value="124"
# sheet.cell(3, 2).value="pooja"
# sheet.cell(3, 3).value="90"
#
# wb.save(file)


#########################################################33
# create excel file

# import openpyxl
# file=r"C:\Users\user\PycharmProjects\pythonProject_WDBatch\selenium\day35\sample1.xlsx"
# wb=openpyxl.Workbook()
# wb.create_sheet("sheet_created")
# print(wb.active)
# sheet = wb["sheet_created"]
# wb.active=sheet
# print(wb.active)
# sheet.cell(1,1).value="data"
# wb.save(file)

import openpyxl
file=r"C:\Users\user\PycharmProjects\pythonProject_WDBatch\selenium\day35\sample2.xlsx"
wb=openpyxl.Workbook()
sheet=wb["Sheet"]
sheet.cell(1,1).value="data"
sheet.title="created_sheet"
wb.save(file)