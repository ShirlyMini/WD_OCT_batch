import openpyxl
from openpyxl.styles import PatternFill

def getRow(xlpath, sheet):
    wb=openpyxl.load_workbook(xlpath)
    sh=wb[sheet]
    return sh.max_row

def getCol(xlpath, sheet):
    wb=openpyxl.load_workbook(xlpath)
    sh=wb[sheet]
    return sh.max_column

def ReadValue(xlpath, sheet, r_no, c_no):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.cell(r_no, c_no).value

def WriteValue(xlpath, sheet, r_no, c_no, data):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    sh.cell(r_no, c_no).value=data
    wb.save(xlpath)

def FillRedColor(xlpath, sheet, r_no, c_no):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    # PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    red = PatternFill(start_color='eb0905', end_color='eb0905', fill_type='solid')
    sh.cell(r_no, c_no).fill = red
    wb.save(xlpath)

def FillGreenColor(xlpath, sheet, r_no, c_no):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    green=PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sh.cell(r_no, c_no).fill = green
    wb.save(xlpath)
